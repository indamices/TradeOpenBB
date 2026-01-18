# Fix typing issues for Python 3.11.0a1
try:
    import fix_typing_notrequired
except ImportError:
    try:
        from . import fix_typing_notrequired
    except ImportError:
        pass

from fastapi import FastAPI, Depends, HTTPException, status, Request, Query
from fastapi.middleware.cors import CORSMiddleware
from fastapi.exceptions import RequestValidationError
from fastapi.responses import JSONResponse, StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy.exc import IntegrityError, SQLAlchemyError
from typing import List, Optional, Dict
import logging
import time

# Use absolute imports for Docker deployment
from database import get_db, init_db
from models import Portfolio, Position, Order, Strategy, AIModelConfig, OrderSide, OrderType, OrderStatus, AIProvider, Base, StockPool, StockInfo, Conversation, ConversationMessage, ChatStrategy, BacktestSymbolList, DataSourceConfig, BacktestRecord
from schemas import (
    Portfolio as PortfolioSchema, PortfolioCreate, PortfolioUpdate,
    Position as PositionSchema, PositionCreate, PositionUpdate,
    Order as OrderSchema, OrderCreate, OrderUpdate,
    Strategy as StrategySchema, StrategyCreate, StrategyUpdate,
    MarketQuote, StrategyGenerationRequest, StrategyGenerationResponse,
    AIModelConfigCreate, AIModelConfigUpdate, AIModelConfigResponse,
    BacktestRequest, BacktestResult, ChatRequest, ChatResponse,
    StockPool as StockPoolSchema, StockPoolCreate, StockPoolUpdate,
    StockInfo as StockInfoSchema, DataSyncRequest, DataSyncResponse,
    Conversation as ConversationSchema, ConversationMessage as ConversationMessageSchema, ChatStrategy as ChatStrategySchema, ChatStrategyCreate, SaveStrategyRequest,
    DataSourceConfigCreate, DataSourceConfigUpdate, DataSourceConfigResponse,
    SymbolList, SymbolListCreate, SymbolListUpdate, SetStrategyActiveRequest, BatchSetActiveRequest,
    BacktestRecord, BacktestRecordCreate, BacktestRecordUpdate,
    ParameterOptimizationRequest, ParameterOptimizationResult
)
from market_service import get_realtime_quote, get_multiple_quotes, get_market_overview, get_technical_indicators
from ai_service_factory import generate_strategy, chat_with_ai
from backtest_engine import run_backtest
from services.benchmark_strategies import list_benchmark_strategies

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Note: Conversation storage is now in database, but keeping this for backward compatibility during migration
conversation_storage: Dict[str, List[Dict]] = {}

app = FastAPI(title="SmartQuant API", version="1.0.0")

# Add rate limiting middleware (300 requests per minute per IP for local development)
# In production, this should be lower (e.g., 60)
try:
    from middleware import RateLimitMiddleware
    import os
    # Higher limit for local development
    rate_limit = int(os.getenv("RATE_LIMIT_PER_MINUTE", "300"))
    app.add_middleware(RateLimitMiddleware, requests_per_minute=rate_limit)
    logger.info(f"Rate limiting middleware enabled: {rate_limit} requests/minute")
except Exception as e:
    logger.warning(f"Failed to enable rate limiting middleware: {str(e)}")

# Global exception handlers
@app.exception_handler(RequestValidationError)
async def validation_exception_handler(request: Request, exc: RequestValidationError):
    """Handle validation errors"""
    logger.warning(f"Validation error: {exc.errors()}")
    return JSONResponse(
        status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
        content={
            "detail": "Validation error",
            "errors": exc.errors()
        }
    )

@app.exception_handler(IntegrityError)
async def integrity_error_handler(request: Request, exc: IntegrityError):
    """Handle database integrity errors"""
    logger.error(f"Database integrity error: {str(exc)}")
    return JSONResponse(
        status_code=status.HTTP_400_BAD_REQUEST,
        content={"detail": "Database integrity error. The operation conflicts with existing data."}
    )

@app.exception_handler(SQLAlchemyError)
async def sqlalchemy_error_handler(request: Request, exc: SQLAlchemyError):
    """Handle SQLAlchemy errors"""
    error_str = str(exc)
    logger.error(f"Database error: {error_str}", exc_info=True)
    
    # Provide more user-friendly error messages for common database errors
    if "could not translate host name" in error_str.lower() or "name or service not known" in error_str.lower():
        detail = (
            "Database connection error: Unable to resolve database hostname. "
            "Please check your DATABASE_URL configuration in Render Dashboard. "
            "Use the External Connection String (with full domain name) instead of Internal. "
            "See RENDER_POSTGRESQL_SETUP.md for detailed instructions."
        )
    elif "connection refused" in error_str.lower() or "connection timeout" in error_str.lower():
        detail = (
            "Database connection error: Unable to connect to database server. "
            "Please check if the database service is running and accessible."
        )
    elif "authentication failed" in error_str.lower() or "password authentication failed" in error_str.lower():
        detail = (
            "Database authentication error: Invalid username or password. "
            "Please check your DATABASE_URL configuration in Render Dashboard."
        )
    elif "database" in error_str.lower() and "does not exist" in error_str.lower():
        detail = (
            "Database error: The specified database does not exist. "
            "Please check your DATABASE_URL configuration."
        )
    else:
        detail = f"Database error occurred: {error_str}"
    
    return JSONResponse(
        status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
        content={"detail": detail}
    )

@app.exception_handler(Exception)
async def general_exception_handler(request: Request, exc: Exception):
    """Handle all other exceptions"""
    logger.error(f"Unhandled exception: {exc}", exc_info=True)
    response = JSONResponse(
        status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
        content={"detail": "Internal server error"}
    )
    # Ensure CORS headers are added even for errors
    origin = request.headers.get("origin")
    if origin:
        # Check if origin is allowed
        allowed_origins = [
            "http://localhost:3000",
            "http://localhost:5173",
            "https://tradeopenbb-frontend.onrender.com",
        ]
        import re
        origin_pattern = re.compile(r"https://.*\.render\.com|https://.*\.railway\.app|https://.*\.fly\.dev|https://.*\.vercel\.app")
        
        if origin in allowed_origins or origin_pattern.match(origin):
            response.headers["Access-Control-Allow-Origin"] = origin
            response.headers["Access-Control-Allow-Credentials"] = "true"
            response.headers["Access-Control-Allow-Methods"] = "*"
            response.headers["Access-Control-Allow-Headers"] = "*"
    return response

# CORS middleware - MUST be added BEFORE RateLimitMiddleware
# In FastAPI, middleware executes in reverse order (last added = first executed)
# So we add CORS last so it wraps all responses, but it will execute first
# Allow local development and cloud platform domains
# Use allow_origin_regex for pattern matching to support wildcard domains
import re

# CORS middleware configuration
# Use CORSMiddleware as the primary CORS handler
# Note: In FastAPI, middleware executes in reverse order (last added = first executed)
# So CORSMiddleware should be added last to execute first and handle CORS properly
# CORS middleware - configure to allow frontend origin
# Note: In production, we need to allow the specific frontend domain
app.add_middleware(
    CORSMiddleware,
    allow_origins=[
        "http://localhost:3000",
        "http://localhost:5173",  # Vite default port
        "https://tradeopenbb-frontend.onrender.com",  # Render frontend
    ],
    allow_origin_regex=r"https://.*\.render\.com|https://.*\.railway\.app|https://.*\.fly\.dev|https://.*\.vercel\.app",
    allow_credentials=True,
    allow_methods=["GET", "POST", "PUT", "DELETE", "OPTIONS", "PATCH", "HEAD"],  # Explicitly list all methods
    allow_headers=["Content-Type", "Authorization", "Accept", "Origin", "X-Requested-With", "X-CSRFToken"],  # Explicit headers
    expose_headers=["*"],
    max_age=3600,  # Cache preflight for 1 hour
)

# Additional CORS ensuring middleware - ensures CORS headers on all responses including errors
# This runs AFTER CORSMiddleware to add headers to error responses that CORSMiddleware might miss
@app.middleware("http")
async def cors_ensuring_middleware(request: Request, call_next):
    """Ensure CORS headers on all responses, including errors"""
    # Let CORSMiddleware handle OPTIONS preflight - don't intercept it here
    response = await call_next(request)
    
    # Ensure CORS headers are present on all responses (including errors that CORSMiddleware might miss)
    origin = request.headers.get("origin")
    if origin and "Access-Control-Allow-Origin" not in response.headers:
        allowed_origins = ["http://localhost:3000", "http://localhost:5173", "https://tradeopenbb-frontend.onrender.com"]
        import re
        origin_pattern = re.compile(r"https://.*\.render\.com|https://.*\.railway\.app|https://.*\.fly\.dev|https://.*\.vercel\.app")
        if origin in allowed_origins or origin_pattern.match(origin):
            response.headers["Access-Control-Allow-Origin"] = origin
            response.headers["Access-Control-Allow-Credentials"] = "true"
            response.headers["Access-Control-Allow-Methods"] = "GET, POST, PUT, DELETE, OPTIONS, PATCH, HEAD"
            response.headers["Access-Control-Allow-Headers"] = "Content-Type, Authorization, Accept, Origin, X-Requested-With, X-CSRFToken"
    
    return response

# Initialize database on startup
@app.on_event("startup")
async def startup_event():
    # Initialize database first
    try:
        from database import init_db
        init_db()
        logger.info("Database initialized successfully")
    except Exception as e:
        logger.error(f"Failed to initialize database: {e}", exc_info=True)
        raise
    
    # Start task scheduler for background data sync
    try:
        from services.scheduler import scheduler
        scheduler.start()
        logger.info("Task scheduler started successfully")
    except Exception as e:
        logger.warning(f"Failed to start task scheduler: {e}. Scheduled tasks will be disabled.")

@app.on_event("shutdown")
async def shutdown_event():
    """Shutdown event handler"""
    try:
        from services.scheduler import scheduler
        scheduler.stop()
        logger.info("Task scheduler stopped")
    except Exception:
        pass
    try:
        # Cleanup on shutdown
        logger.info("Application shutting down")
        
        # Verify default portfolio exists
        db = next(get_db())
        try:
            portfolio = db.query(Portfolio).filter(Portfolio.id == 1).first()
            if portfolio:
                logger.info(f"Verified default portfolio exists: {portfolio.name} (ID: {portfolio.id})")
            else:
                logger.warning("Default portfolio (ID=1) not found after initialization")
        finally:
            db.close()
    except Exception as e:
        logger.error(f"Database initialization failed: {str(e)}", exc_info=True)
        # Don't raise - allow app to start even if DB init fails

# Health check
@app.get("/")
@app.get("/health")
async def root():
    """Health check endpoint - optimized for fast response"""
    return {"message": "SmartQuant API", "status": "running", "timestamp": time.time()}

# Portfolio endpoints
@app.get("/api/portfolio", response_model=PortfolioSchema)
async def get_portfolio(portfolio_id: int = 1, db: Session = Depends(get_db)):
    portfolio = db.query(Portfolio).filter(Portfolio.id == portfolio_id).first()
    if not portfolio:
        raise HTTPException(status_code=404, detail="Portfolio not found")
    return portfolio

@app.post("/api/portfolio", response_model=PortfolioSchema, status_code=status.HTTP_201_CREATED)
async def create_portfolio(portfolio: PortfolioCreate, db: Session = Depends(get_db)):
    db_portfolio = Portfolio(
        name=portfolio.name,
        initial_cash=portfolio.initial_cash,
        current_cash=portfolio.initial_cash,
        total_value=portfolio.initial_cash
    )
    db.add(db_portfolio)
    db.commit()
    db.refresh(db_portfolio)
    return db_portfolio

@app.put("/api/portfolio/{portfolio_id}", response_model=PortfolioSchema)
async def update_portfolio(portfolio_id: int, portfolio: PortfolioUpdate, db: Session = Depends(get_db)):
    db_portfolio = db.query(Portfolio).filter(Portfolio.id == portfolio_id).first()
    if not db_portfolio:
        raise HTTPException(status_code=404, detail="Portfolio not found")
    
    update_data = portfolio.model_dump(exclude_unset=True)
    for field, value in update_data.items():
        setattr(db_portfolio, field, value)
    
    db.commit()
    db.refresh(db_portfolio)
    return db_portfolio

# Position endpoints
@app.get("/api/positions", response_model=List[PositionSchema])
async def get_positions(portfolio_id: int = 1, db: Session = Depends(get_db)):
    positions = db.query(Position).filter(Position.portfolio_id == portfolio_id).all()
    return positions

@app.post("/api/positions", response_model=PositionSchema, status_code=status.HTTP_201_CREATED)
async def create_position(position: PositionCreate, db: Session = Depends(get_db)):
    # Verify portfolio exists
    portfolio = db.query(Portfolio).filter(Portfolio.id == position.portfolio_id).first()
    if not portfolio:
        raise HTTPException(status_code=404, detail="Portfolio not found")
    
    position_dict = position.model_dump()
    # Calculate market_value if not provided
    if 'market_value' not in position_dict:
        position_dict['market_value'] = position_dict['quantity'] * position_dict['current_price']
    # Calculate unrealized P&L
    if 'unrealized_pnl' not in position_dict:
        cost = position_dict['quantity'] * position_dict['avg_price']
        market_value = position_dict['market_value']
        position_dict['unrealized_pnl'] = market_value - cost
    if 'unrealized_pnl_percent' not in position_dict:
        cost = position_dict['quantity'] * position_dict['avg_price']
        if cost > 0:
            position_dict['unrealized_pnl_percent'] = ((position_dict['current_price'] - position_dict['avg_price']) / position_dict['avg_price']) * 100
        else:
            position_dict['unrealized_pnl_percent'] = 0.0
    
    db_position = Position(**position_dict)
    db.add(db_position)
    db.commit()
    db.refresh(db_position)
    return db_position

@app.put("/api/positions/{position_id}", response_model=PositionSchema)
async def update_position(position_id: int, position: PositionUpdate, db: Session = Depends(get_db)):
    """Update a position"""
    db_position = db.query(Position).filter(Position.id == position_id).first()
    if not db_position:
        raise HTTPException(status_code=404, detail="Position not found")
    
    update_data = position.model_dump(exclude_unset=True)
    
    # Recalculate market_value and P&L if price or quantity changed
    if 'current_price' in update_data or 'quantity' in update_data:
        current_price = update_data.get('current_price', db_position.current_price)
        quantity = update_data.get('quantity', db_position.quantity)
        avg_price = update_data.get('avg_price', db_position.avg_price)
        
        update_data['market_value'] = quantity * current_price
        cost = quantity * avg_price
        update_data['unrealized_pnl'] = update_data['market_value'] - cost
        if cost > 0:
            update_data['unrealized_pnl_percent'] = ((current_price - avg_price) / avg_price) * 100
        else:
            update_data['unrealized_pnl_percent'] = 0.0
    
    for field, value in update_data.items():
        setattr(db_position, field, value)
    
    db.commit()
    db.refresh(db_position)
    return db_position

@app.delete("/api/positions/{position_id}", status_code=status.HTTP_204_NO_CONTENT)
async def delete_position(position_id: int, db: Session = Depends(get_db)):
    """Delete a position"""
    db_position = db.query(Position).filter(Position.id == position_id).first()
    if not db_position:
        raise HTTPException(status_code=404, detail="Position not found")
    
    db.delete(db_position)
    db.commit()
    return None

# Order endpoints
@app.get("/api/orders", response_model=List[OrderSchema])
async def get_orders(
    portfolio_id: int = 1,
    skip: int = 0,
    limit: int = 100,
    db: Session = Depends(get_db)
):
    """Get orders with pagination"""
    orders = db.query(Order)\
        .filter(Order.portfolio_id == portfolio_id)\
        .order_by(Order.created_at.desc())\
        .offset(skip)\
        .limit(limit)\
        .all()
    return orders

@app.post("/api/orders", response_model=OrderSchema, status_code=status.HTTP_201_CREATED)
async def create_order(order: OrderCreate, db: Session = Depends(get_db)):
    # Verify portfolio exists
    portfolio = db.query(Portfolio).filter(Portfolio.id == order.portfolio_id).first()
    if not portfolio:
        raise HTTPException(status_code=404, detail="Portfolio not found")
    
    db_order = Order(**order.model_dump(), status=OrderStatus.PENDING)
    db.add(db_order)
    db.commit()
    db.refresh(db_order)
    return db_order

# Strategy endpoints
@app.get("/api/strategies", response_model=List[StrategySchema])
async def get_strategies(
    portfolio_id: Optional[int] = None,
    is_active: Optional[bool] = None,  # 新增：按活跃状态过滤
    db: Session = Depends(get_db)
):
    """获取策略列表（支持按活跃状态过滤）"""
    query = db.query(Strategy)
    if portfolio_id:
        query = query.filter(Strategy.target_portfolio_id == portfolio_id)
    if is_active is not None:
        query = query.filter(Strategy.is_active == is_active)
    strategies = query.order_by(Strategy.created_at.desc()).all()
    return strategies

@app.get("/api/strategies/active", response_model=List[StrategySchema])
async def get_active_strategies(
    portfolio_id: Optional[int] = None,
    db: Session = Depends(get_db)
):
    """获取所有活跃策略（用于回测实验室）"""
    query = db.query(Strategy).filter(Strategy.is_active == True)
    if portfolio_id:
        query = query.filter(Strategy.target_portfolio_id == portfolio_id)
    strategies = query.order_by(Strategy.created_at.desc()).all()
    return strategies

@app.post("/api/strategies", response_model=StrategySchema, status_code=status.HTTP_201_CREATED)
async def create_strategy(strategy: StrategyCreate, db: Session = Depends(get_db)):
    db_strategy = Strategy(**strategy.model_dump())
    db.add(db_strategy)
    db.commit()
    db.refresh(db_strategy)
    return db_strategy

@app.put("/api/strategies/{strategy_id}", response_model=StrategySchema)
async def update_strategy(strategy_id: int, strategy: StrategyUpdate, db: Session = Depends(get_db)):
    """更新策略（包括活跃状态）"""
    db_strategy = db.query(Strategy).filter(Strategy.id == strategy_id).first()
    if not db_strategy:
        raise HTTPException(status_code=404, detail="Strategy not found")
    
    update_data = strategy.model_dump(exclude_unset=True)
    for field, value in update_data.items():
        setattr(db_strategy, field, value)
    
    db.commit()
    db.refresh(db_strategy)
    return db_strategy

@app.put("/api/strategies/{strategy_id}/set-active", response_model=StrategySchema)
async def set_strategy_active(
    strategy_id: int,
    request: SetStrategyActiveRequest,
    db: Session = Depends(get_db)
):
    """设置策略活跃状态（显式设置True/False）"""
    db_strategy = db.query(Strategy).filter(Strategy.id == strategy_id).first()
    if not db_strategy:
        raise HTTPException(status_code=404, detail="Strategy not found")
    
    db_strategy.is_active = request.is_active
    db.commit()
    db.refresh(db_strategy)
    return db_strategy

@app.put("/api/strategies/{strategy_id}/toggle-active", response_model=StrategySchema)
async def toggle_strategy_active(strategy_id: int, db: Session = Depends(get_db)):
    """切换策略活跃状态（当前状态取反）"""
    db_strategy = db.query(Strategy).filter(Strategy.id == strategy_id).first()
    if not db_strategy:
        raise HTTPException(status_code=404, detail="Strategy not found")
    
    db_strategy.is_active = not db_strategy.is_active
    db.commit()
    db.refresh(db_strategy)
    return db_strategy

@app.put("/api/strategies/batch-set-active")
async def batch_set_strategy_active(
    request: BatchSetActiveRequest,
    db: Session = Depends(get_db)
):
    """批量设置多个策略的活跃状态"""
    strategies = db.query(Strategy).filter(Strategy.id.in_(request.strategy_ids)).all()
    if len(strategies) != len(request.strategy_ids):
        raise HTTPException(status_code=404, detail="Some strategies not found")
    
    for strategy in strategies:
        strategy.is_active = request.is_active
    
    db.commit()
    return {"updated": len(strategies), "strategies": [s.id for s in strategies]}

@app.delete("/api/strategies/{strategy_id}", status_code=status.HTTP_204_NO_CONTENT)
async def delete_strategy(strategy_id: int, db: Session = Depends(get_db)):
    """删除策略"""
    db_strategy = db.query(Strategy).filter(Strategy.id == strategy_id).first()
    if not db_strategy:
        raise HTTPException(status_code=404, detail="Strategy not found")
    
    # 更新关联的ChatStrategy记录
    db.query(ChatStrategy).filter(ChatStrategy.saved_strategy_id == strategy_id).update({
        "is_saved": False,
        "saved_strategy_id": None
    })
    
    db.delete(db_strategy)
    db.commit()
    return None

@app.post("/api/strategies/generate", response_model=StrategyGenerationResponse)
async def generate_strategy_endpoint(request: StrategyGenerationRequest, db: Session = Depends(get_db)):
    """Generate strategy code using AI"""
    try:
        result = await generate_strategy(request.prompt, request.model_id, db)
        return result
    except Exception as e:
        logger.error(f"Strategy generation failed: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Strategy generation failed: {str(e)}")

# AI Chat endpoints (使用数据库持久化)
@app.post("/api/ai/chat", response_model=ChatResponse)
async def chat_endpoint(request: ChatRequest, db: Session = Depends(get_db)):
    """AI chat conversation (持久化到数据库)"""
    try:
        import uuid
        from datetime import datetime
        
        # Get or create conversation
        conversation_id = request.conversation_id
        conversation = None
        
        if conversation_id:
            conversation = db.query(Conversation).filter(
                Conversation.conversation_id == conversation_id
            ).first()
        
        # Create new conversation if needed
        if not conversation:
            conversation_id = str(uuid.uuid4())
            # 从第一条消息提取标题（前50个字符）
            title = request.message[:50] if request.message else "New Conversation"
            conversation = Conversation(
                conversation_id=conversation_id,
                title=title
            )
            db.add(conversation)
            db.commit()
            db.refresh(conversation)
        
        # Save user message to database
        user_message = ConversationMessage(
            conversation_id=conversation_id,
            role='user',
            content=request.message
        )
        db.add(user_message)
        db.commit()
        
        # Get conversation history for AI context
        messages = db.query(ConversationMessage).filter(
            ConversationMessage.conversation_id == conversation_id
        ).order_by(ConversationMessage.created_at).all()
        
        # Convert to format expected by AI service
        conversation_history = [
            {
                'role': msg.role,
                'content': msg.content,
                'timestamp': msg.created_at.isoformat()
            }
            for msg in messages[:-1]  # Exclude current message
        ]
        
        # Call AI service
        try:
            ai_response = await chat_with_ai(
                request.message,
                None,  # Use default model
                db,
                conversation_history=conversation_history
            )
            code_snippets = None
            
            # 尝试从响应中提取代码片段
            import re
            code_pattern = r'```python\n(.*?)\n```'
            matches = re.findall(code_pattern, ai_response, re.DOTALL)
            if matches:
                code_snippets = {"python": matches[0]}
                logger.info(f"Extracted code snippet from AI response")
            
        except Exception as e:
            logger.error(f"Chat failed: {str(e)}", exc_info=True)
            ai_response = "抱歉，发生了错误。请重试或检查AI模型配置。"
            code_snippets = None
        
        # Save assistant response to database
        assistant_message = ConversationMessage(
            conversation_id=conversation_id,
            role='assistant',
            content=ai_response,
            code_snippets=code_snippets
        )
        db.add(assistant_message)
        
        # Update conversation updated_at
        conversation.updated_at = datetime.now()
        
        db.commit()
        
        return ChatResponse(
            message=ai_response,
            conversation_id=conversation_id,
            suggestions=None,
            code_snippets=code_snippets
        )
    except Exception as e:
        logger.error(f"Chat failed: {str(e)}", exc_info=True)
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Chat failed: {str(e)}")

@app.get("/api/ai/conversations", response_model=List[ConversationSchema])
async def get_conversations(db: Session = Depends(get_db)):
    """获取所有聊天会话列表"""
    try:
        conversations = db.query(Conversation).order_by(
            Conversation.updated_at.desc()
        ).all()
        
        # 添加消息数量
        result = []
        for conv in conversations:
            message_count = db.query(ConversationMessage).filter(
                ConversationMessage.conversation_id == conv.conversation_id
            ).count()
            
            conv_dict = {
                "id": conv.id,
                "conversation_id": conv.conversation_id,
                "title": conv.title,
                "created_at": conv.created_at,
                "updated_at": conv.updated_at,
                "message_count": message_count
            }
            result.append(conv_dict)
        
        return result
    except Exception as e:
        logger.error(f"Failed to get conversations: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to get conversations: {str(e)}")

@app.get("/api/ai/conversations/{conversation_id}")
async def get_conversation(conversation_id: str, db: Session = Depends(get_db)):
    """获取会话详情（包含消息）"""
    try:
        conversation = db.query(Conversation).filter(
            Conversation.conversation_id == conversation_id
        ).first()
        
        if not conversation:
            raise HTTPException(status_code=404, detail="Conversation not found")
        
        # Get messages
        messages = db.query(ConversationMessage).filter(
            ConversationMessage.conversation_id == conversation_id
        ).order_by(ConversationMessage.created_at).all()
        
        # Convert to ChatMessage format
        chat_messages = [
            {
                "id": msg.id,
                "conversation_id": msg.conversation_id,
                "role": msg.role,
                "content": msg.content,
                "code_snippets": msg.code_snippets,
                "timestamp": msg.created_at.isoformat()
            }
            for msg in messages
        ]
        
        return {
            "conversation_id": conversation.conversation_id,
            "title": conversation.title,
            "created_at": conversation.created_at.isoformat(),
            "updated_at": conversation.updated_at.isoformat() if conversation.updated_at else None,
            "messages": chat_messages
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Failed to get conversation: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to get conversation: {str(e)}")

@app.delete("/api/ai/conversations/{conversation_id}", status_code=status.HTTP_204_NO_CONTENT)
async def delete_conversation(conversation_id: str, db: Session = Depends(get_db)):
    """删除会话及其消息"""
    try:
        conversation = db.query(Conversation).filter(
            Conversation.conversation_id == conversation_id
        ).first()
        
        if not conversation:
            raise HTTPException(status_code=404, detail="Conversation not found")
        
        # Messages will be deleted via cascade
        db.delete(conversation)
        db.commit()
        return None
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Failed to delete conversation: {str(e)}")
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Failed to delete conversation: {str(e)}")

@app.get("/api/ai/chat/{conversation_id}")
async def get_conversation_history(conversation_id: str, db: Session = Depends(get_db)):
    """Get conversation history (兼容旧API)"""
    try:
        conversation = db.query(Conversation).filter(
            Conversation.conversation_id == conversation_id
        ).first()
        
        if not conversation:
            return {"conversation_id": conversation_id, "messages": []}
        
        messages = db.query(ConversationMessage).filter(
            ConversationMessage.conversation_id == conversation_id
        ).order_by(ConversationMessage.created_at).all()
        
        # Convert to old format for compatibility
        message_list = [
            {
                'role': msg.role,
                'content': msg.content,
                'timestamp': msg.created_at.isoformat(),
                'code_snippets': msg.code_snippets
            }
            for msg in messages
        ]
        
        return {
            "conversation_id": conversation_id,
            "messages": message_list
        }
    except Exception as e:
        logger.error(f"Failed to get conversation history: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to get conversation history: {str(e)}")

@app.post("/api/ai/suggestions")
async def get_suggestions(context: Optional[Dict] = None, db: Session = Depends(get_db)):
    """Get AI proactive strategy suggestions"""
    try:
        # Simplified suggestions - can be enhanced with more sophisticated logic
        suggestions = [
            "Consider a mean reversion strategy using Bollinger Bands",
            "Try a momentum strategy based on RSI indicators",
            "Explore a pairs trading strategy for correlated stocks"
        ]
        return {"suggestions": suggestions}
    except Exception as e:
        logger.error(f"Failed to get suggestions: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to get suggestions: {str(e)}")

# Market endpoints
@app.get("/api/market/quote/{symbol}", response_model=MarketQuote)
async def get_quote(symbol: str):
    """Get real-time market quote"""
    try:
        quote = await get_realtime_quote(symbol.upper())
        return quote
    except Exception as e:
        logger.error(f"Failed to get quote for {symbol}: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to get quote: {str(e)}")

@app.get("/api/market/quotes", response_model=List[MarketQuote])
async def get_quotes(symbols: str):
    """
    Get real-time quotes for multiple symbols
    Query parameter: symbols (comma-separated, e.g., "AAPL,MSFT,GOOGL")
    Limited to 20 symbols per request to prevent overload
    """
    try:
        symbol_list = [s.strip().upper() for s in symbols.split(',')]
        
        # Limit number of symbols per request
        if len(symbol_list) > 20:
            logger.warning(f"Too many symbols requested ({len(symbol_list)}), limiting to 20")
            symbol_list = symbol_list[:20]
        
        quotes = await get_multiple_quotes(symbol_list)
        return quotes
    except Exception as e:
        logger.error(f"Failed to get quotes: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to get quotes: {str(e)}")

@app.get("/api/market/indicators/{symbol}")
async def get_indicators(symbol: str, indicators: str = "MACD,RSI,BB", period: int = 20):
    """
    Get technical indicators for a symbol
    Query parameters:
    - indicators: comma-separated indicator names (default: "MACD,RSI,BB")
    - period: period for calculations (default: 20)
    """
    try:
        indicator_list = [i.strip() for i in indicators.split(',')]
        data = await get_technical_indicators(symbol.upper(), indicator_list, period)
        # Convert DataFrame to dict for JSON response
        import numpy as np
        import pandas as pd
        
        if isinstance(data, dict):
            # Handle dict case - recursively clean NaN/Inf values
            def clean_dict(d):
                if isinstance(d, dict):
                    return {k: clean_dict(v) for k, v in d.items()}
                elif isinstance(d, list):
                    return [clean_dict(item) for item in d]
                elif isinstance(d, (float, np.floating)):
                    if np.isnan(d) or np.isinf(d):
                        return None
                    return float(d)
                return d
            return clean_dict(data)
        elif hasattr(data, 'to_dict'):
            # DataFrame case
            data_clean = data.replace([np.inf, -np.inf], np.nan)
            data_clean = data_clean.where(pd.notnull(data_clean), None)
            result_dict = data_clean.to_dict(orient='records')
            # Additional cleaning pass on the dict result using sanitize_for_json
            from utils.json_serializer import sanitize_for_json
            result = sanitize_for_json(result_dict)
            return result
        else:
            # Other types - try to convert to JSON-serializable format
            return data
    except Exception as e:
        logger.error(f"Failed to get indicators for {symbol}: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to get indicators: {str(e)}")

@app.get("/api/market/overview")
async def get_overview():
    """Get market overview data (cached for 30 seconds)"""
    try:
        # Use cache to reduce API calls
        from cachetools import TTLCache
        cache_key = "market_overview"
        
        # Environment-based cache configuration
        # Production: shorter cache for data freshness (30 seconds)
        # Development: longer cache to reduce API calls (60 seconds)
        ENVIRONMENT = os.getenv("ENVIRONMENT", "development")
        cache_ttl = 30 if ENVIRONMENT == "production" else 60
        
        if not hasattr(get_overview, '_cache'):
            get_overview._cache = TTLCache(maxsize=1, ttl=cache_ttl)
        
        if cache_key in get_overview._cache:
            return get_overview._cache[cache_key]
        
        overview = await get_market_overview()
        get_overview._cache[cache_key] = overview
        return overview
    except Exception as e:
        logger.error(f"Failed to get market overview: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to get market overview: {str(e)}")

@app.get("/api/market/historical/{symbol}")
async def get_historical_market_data(
    symbol: str,
    start_date: str = Query(..., description="Start date in YYYY-MM-DD format"),
    end_date: str = Query(..., description="End date in YYYY-MM-DD format"),
    db: Session = Depends(get_db)
):
    """
    Get historical market data for a symbol
    
    Query parameters:
    - symbol: Stock symbol (e.g., 'AAPL')
    - start_date: Start date in 'YYYY-MM-DD' format
    - end_date: End date in 'YYYY-MM-DD' format
    """
    try:
        from services.data_service import DataService
        
        async with DataService(db=db) as data_service:
            data = await data_service.get_historical_data(
                symbol.upper(),
                start_date,
                end_date,
                use_cache=True
            )
            
            if data is None or data.empty:
                raise HTTPException(
                    status_code=404, 
                    detail=f"No historical data found for {symbol} in the specified date range"
                )
            
            # Convert DataFrame to JSON-serializable format
            import numpy as np
            import pandas as pd
            from utils.json_serializer import sanitize_for_json
            
            # Reset index to make Date a column
            data_reset = data.reset_index()
            
            # Clean data
            data_clean = data_reset.replace([np.inf, -np.inf], np.nan)
            data_clean = data_clean.where(pd.notnull(data_clean), None)
            
            # Convert to dict
            result = data_clean.to_dict(orient='records')
            
            # Sanitize for JSON
            result = sanitize_for_json(result)
            
            return result
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Failed to get historical data for {symbol}: {str(e)}")
        raise HTTPException(
            status_code=500, 
            detail=f"Failed to get historical data: {str(e)}"
        )

# Backtest endpoints
@app.post("/api/backtest", response_model=BacktestResult)
async def run_backtest_endpoint(
    request: BacktestRequest, 
    db: Session = Depends(get_db),
    save_record: bool = Query(False, description="是否保存回测记录")
):
    """Run backtest for a strategy"""
    try:
        result = await run_backtest(request, db)
        
        # Add index comparison if requested (legacy support)
        if request.compare_with_indices:
            try:
                from services.index_comparison import compare_with_indices
                comparisons = await compare_with_indices(
                    result.model_dump(),
                    request.start_date,
                    request.end_date
                )
                result_dict = result.model_dump()
                result_dict['index_comparisons'] = comparisons
                from schemas import BacktestResult
                result = BacktestResult(**result_dict)
            except Exception as e:
                logger.warning(f"Index comparison failed: {str(e)}")
        
        # Add strategy comparison if compare_items is provided
        if request.compare_items and len(request.compare_items) > 0:
            try:
                from services.strategy_comparison import compare_strategies
                from services.data_service import DataService
                from database import SessionLocal
                
                # Get historical data (reuse from backtest if possible, otherwise fetch)
                all_data = {}
                try:
                    db_data = SessionLocal()
                    try:
                        async with DataService(db=db_data) as data_service:
                            data_dict = await data_service.batch_fetch_historical_data(
                                symbols=request.symbols,
                                start_date=request.start_date,
                                end_date=request.end_date
                            )
                            all_data = data_dict
                    finally:
                        if db_data:
                            db_data.close()
                except Exception as e:
                    logger.warning(f"Failed to fetch data for comparison: {str(e)}")
                
                if all_data:
                    comparisons = await compare_strategies(
                        main_result=result,
                        request=request,
                        compare_items=request.compare_items,
                        all_data=all_data,
                        db=db
                    )
                    result_dict = result.model_dump()
                    result_dict['strategy_comparisons'] = comparisons
                    from schemas import BacktestResult
                    result = BacktestResult(**result_dict)
            except Exception as e:
                logger.warning(f"Strategy comparison failed: {str(e)}")
        
        # If requested, save backtest record
        if save_record:
            try:
                strategy = db.query(Strategy).filter(Strategy.id == request.strategy_id).first()
                if strategy:
                    from datetime import datetime as dt
                    
                    # Convert date strings to date objects
                    start_dt = dt.strptime(request.start_date, '%Y-%m-%d').date()
                    end_dt = dt.strptime(request.end_date, '%Y-%m-%d').date()
                    
                    backtest_record = BacktestRecord(
                        strategy_id=request.strategy_id,
                        strategy_name=strategy.name,
                        start_date=start_dt,
                        end_date=end_dt,
                        initial_cash=request.initial_cash,
                        symbols=request.symbols,
                        sharpe_ratio=result.sharpe_ratio,
                        sortino_ratio=result.sortino_ratio,
                        annualized_return=result.annualized_return,
                        max_drawdown=result.max_drawdown,
                        win_rate=result.win_rate,
                        total_trades=result.total_trades,
                        total_return=result.total_return,
                        compare_with_indices=request.compare_with_indices or False,
                        compare_items=request.compare_items,
                        full_result=result.model_dump()  # Save complete result as JSON
                    )
                    
                    db.add(backtest_record)
                    db.commit()
                    db.refresh(backtest_record)
                    logger.info(f"Backtest record saved with ID: {backtest_record.id}")
            except Exception as e:
                logger.error(f"Failed to save backtest record: {str(e)}")
                # Don't interrupt the backtest flow, just log the error
        
        return result
    except Exception as e:
        logger.error(f"Backtest failed: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Backtest failed: {str(e)}")

@app.post("/api/backtest/optimize", response_model=ParameterOptimizationResult)
async def optimize_strategy_parameters(
    request: ParameterOptimizationRequest,
    db: Session = Depends(get_db)
):
    """Optimize strategy parameters using grid search"""
    try:
        from services.parameter_optimizer import ParameterOptimizer
        
        optimizer = ParameterOptimizer(db)
        optimization_result = await optimizer.optimize_parameters(
            strategy_id=request.strategy_id,
            optimization_request=request,
            optimization_metric=request.optimization_metric
        )
        
        # Convert result format to match schema
        # Backend returns: {best_parameters, best_metric_value, best_result, all_results, ...}
        # Schema expects: {best_parameters, best_metric_value, optimization_metric, results, total_combinations}
        formatted_results = []
        for item in optimization_result.get('all_results', []):
            if 'error' not in item:
                # Extract metrics from result dict
                result_dict = item.get('result', {})
                formatted_results.append({
                    'parameters': item.get('parameters', {}),
                    'metrics': result_dict,  # Full backtest result
                    # Also include metrics at top level for compatibility
                    'sharpe_ratio': result_dict.get('sharpe_ratio'),
                    'sortino_ratio': result_dict.get('sortino_ratio'),
                    'annualized_return': result_dict.get('annualized_return'),
                    'total_return': result_dict.get('total_return'),
                    'max_drawdown': result_dict.get('max_drawdown'),
                    'win_rate': result_dict.get('win_rate'),
                    'total_trades': result_dict.get('total_trades')
                })
        
        return ParameterOptimizationResult(
            best_parameters=optimization_result.get('best_parameters', {}),
            best_metric_value=optimization_result.get('best_metric_value', 0),
            optimization_metric=optimization_result.get('optimization_metric', 'sharpe_ratio'),
            results=formatted_results,
            total_combinations=optimization_result.get('total_combinations', 0)
        )
    except Exception as e:
        logger.error(f"Parameter optimization failed: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Parameter optimization failed: {str(e)}")

@app.post("/api/backtest/analyze", response_model=AIStrategyAnalysisResponse)
async def analyze_backtest_result(
    request: AIStrategyAnalysisRequest,
    db: Session = Depends(get_db)
):
    """Analyze backtest result using AI and provide suggestions"""
    try:
        from services.strategy_analyzer import StrategyAnalyzer
        from models import Strategy
        
        # Get strategy
        strategy = db.query(Strategy).filter(Strategy.id == request.strategy_id).first()
        if not strategy:
            raise HTTPException(status_code=404, detail="Strategy not found")
        
        analyzer = StrategyAnalyzer(db)
        
        # Convert backtest_result to dict if it's a Pydantic model
        backtest_result_dict = request.backtest_result
        if hasattr(request.backtest_result, 'model_dump'):
            backtest_result_dict = request.backtest_result.model_dump()
        elif hasattr(request.backtest_result, '__dict__'):
            backtest_result_dict = request.backtest_result.__dict__
        elif isinstance(request.backtest_result, dict):
            backtest_result_dict = request.backtest_result
        
        result = await analyzer.analyze_backtest_result(
            backtest_result=backtest_result_dict,
            strategy_code=strategy.logic_code,
            strategy_name=strategy.name
        )
        
        return AIStrategyAnalysisResponse(**result)
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Strategy analysis failed: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Strategy analysis failed: {str(e)}")

# AI Model Config endpoints
@app.get("/api/ai-models", response_model=List[AIModelConfigResponse])
async def get_ai_models(db: Session = Depends(get_db)):
    """Get all AI model configurations"""
    try:
        # Show all models (including inactive ones) so users can activate/deactivate them
        models = db.query(AIModelConfig).order_by(AIModelConfig.is_active.desc(), AIModelConfig.is_default.desc(), AIModelConfig.created_at.desc()).all()
        result = []
        for model in models:
            # Handle provider enum conversion - SQLAlchemy may return string or enum
            # This handles the case where database stores 'custom' but SQLAlchemy expects CUSTOM
            provider_value = None
            try:
                if isinstance(model.provider, AIProvider):
                    # Already an enum, get its value
                    provider_value = model.provider.value
                elif isinstance(model.provider, str):
                    # It's a string from database, validate and use it
                    # Try to convert to enum first to validate
                    try:
                        provider_enum = AIProvider(model.provider.lower())
                        provider_value = provider_enum.value
                    except (ValueError, AttributeError):
                        # If conversion fails, use the string as-is (for backward compatibility)
                        logger.warning(f"Invalid provider value '{model.provider}' for model {model.id}, using as-is")
                        provider_value = model.provider.lower()
                elif hasattr(model.provider, 'value'):
                    # Has value attribute (enum-like object)
                    provider_value = model.provider.value
                else:
                    # Fallback: convert to string and lowercase
                    provider_str = str(model.provider)
                    try:
                        provider_enum = AIProvider(provider_str.lower())
                        provider_value = provider_enum.value
                    except (ValueError, AttributeError):
                        provider_value = provider_str.lower()
            except Exception as e:
                logger.warning(f"Error processing provider for model {model.id}: {str(e)}")
                # Fallback to string representation
                provider_value = str(model.provider).lower() if model.provider else 'custom'
            
            result.append(AIModelConfigResponse(
                id=model.id,
                name=model.name,
                provider=provider_value,
                model_name=model.model_name,
                base_url=model.base_url,
                is_default=model.is_default,
                is_active=model.is_active,
                created_at=model.created_at
            ))
        return result
    except Exception as e:
        logger.error(f"Error fetching AI models: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Failed to fetch AI models: {str(e)}")

@app.post("/api/ai-models", response_model=AIModelConfigResponse, status_code=status.HTTP_201_CREATED)
async def create_ai_model(model: AIModelConfigCreate, db: Session = Depends(get_db)):
    """Create new AI model configuration"""
    try:
        from ai_service_factory import encrypt_api_key
        
        logger.info(f"Creating AI model: {model.name}, provider: {model.provider}")
        
        # Validate provider enum - handle both string and enum types
        try:
            if isinstance(model.provider, str):
                # Convert string to enum
                provider_enum = AIProvider(model.provider.lower())
            elif isinstance(model.provider, AIProvider):
                provider_enum = model.provider
            else:
                # Try to get value if it's an enum with value attribute
                provider_value = getattr(model.provider, 'value', str(model.provider))
                provider_enum = AIProvider(provider_value.lower())
        except (ValueError, AttributeError) as e:
            logger.error(f"Invalid provider: {model.provider}, type: {type(model.provider)}, error: {str(e)}")
            valid_providers = [p.value for p in AIProvider]
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"Invalid provider: {model.provider}. Must be one of: {valid_providers}"
            )
        
        # Encrypt API key
        try:
            encrypted_key = encrypt_api_key(model.api_key)
            logger.info("API key encrypted successfully")
        except Exception as e:
            logger.error(f"Failed to encrypt API key: {str(e)}", exc_info=True)
            raise HTTPException(
                status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                detail=f"Failed to encrypt API key: {str(e)}"
            )
        
        # Create database model
        db_model = AIModelConfig(
            name=model.name,
            provider=provider_enum,
            api_key=encrypted_key,
            base_url=model.base_url,
            model_name=model.model_name,
            is_default=False,
            is_active=True
        )
        
        # If this is set as default, unset others
        if hasattr(model, 'is_default') and model.is_default:
            db.query(AIModelConfig).update({"is_default": False})
            db_model.is_default = True
        
        db.add(db_model)
        db.commit()
        db.refresh(db_model)
        
        logger.info(f"AI model created successfully with ID: {db_model.id}")
        
        # Return response with proper serialization
        return {
            "id": db_model.id,
            "name": db_model.name,
            "provider": db_model.provider.value if hasattr(db_model.provider, 'value') else str(db_model.provider),
            "model_name": db_model.model_name,
            "base_url": db_model.base_url,
            "is_default": db_model.is_default,
            "is_active": db_model.is_active,
            "created_at": db_model.created_at
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error creating AI model: {str(e)}", exc_info=True)
        db.rollback()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to create AI model: {str(e)}"
        )

@app.put("/api/ai-models/{model_id}", response_model=AIModelConfigResponse)
async def update_ai_model(model_id: int, model: AIModelConfigUpdate, db: Session = Depends(get_db)):
    """Update AI model configuration"""
    from ai_service_factory import encrypt_api_key
    
    db_model = db.query(AIModelConfig).filter(AIModelConfig.id == model_id).first()
    if not db_model:
        raise HTTPException(status_code=404, detail="AI model not found")
    
    update_data = model.model_dump(exclude_unset=True)
    
    # CRITICAL FIX: Only encrypt and update API key if a new one is provided
    # If api_key is empty or not provided, keep the existing encrypted key
    if "api_key" in update_data:
        if update_data["api_key"] and update_data["api_key"].strip():
            # New API key provided, encrypt it
            try:
                update_data["api_key"] = encrypt_api_key(update_data["api_key"].strip())
                logger.info(f"API key updated for model {model_id}")
            except Exception as e:
                logger.error(f"Failed to encrypt API key: {str(e)}")
                raise HTTPException(
                    status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                    detail=f"Failed to encrypt API key: {str(e)}"
                )
        else:
            # Empty API key provided, remove it from update_data to keep existing key
            logger.info(f"Empty API key provided for model {model_id}, keeping existing key")
            del update_data["api_key"]
    
    for field, value in update_data.items():
        setattr(db_model, field, value)
    
    db.commit()
    db.refresh(db_model)
    
    # Return response with proper serialization
    provider_value = db_model.provider.value if hasattr(db_model.provider, 'value') else str(db_model.provider)
    return {
        "id": db_model.id,
        "name": db_model.name,
        "provider": provider_value,
        "model_name": db_model.model_name,
        "base_url": db_model.base_url,
        "is_default": db_model.is_default,
        "is_active": db_model.is_active,
        "created_at": db_model.created_at
    }

@app.delete("/api/ai-models/{model_id}", status_code=status.HTTP_204_NO_CONTENT)
async def delete_ai_model(model_id: int, db: Session = Depends(get_db)):
    """Delete AI model configuration"""
    db_model = db.query(AIModelConfig).filter(AIModelConfig.id == model_id).first()
    if not db_model:
        raise HTTPException(status_code=404, detail="AI model not found")
    
    db.delete(db_model)
    db.commit()
    return None

@app.post("/api/ai-models/{model_id}/test")
async def test_ai_model(model_id: int, db: Session = Depends(get_db)):
    """Test AI model connection"""
    from ai_service_factory import check_ai_model_connection
    
    try:
        result = await check_ai_model_connection(model_id, db)
        return {"success": True, "message": "Model connection successful"}
    except Exception as e:
        return {"success": False, "message": str(e)}

@app.put("/api/ai-models/{model_id}/set-default", response_model=AIModelConfigResponse)
async def set_default_ai_model(model_id: int, db: Session = Depends(get_db)):
    """Set default AI model (for backward compatibility)"""
    db_model = db.query(AIModelConfig).filter(AIModelConfig.id == model_id).first()
    if not db_model:
        raise HTTPException(status_code=404, detail="AI model not found")
    
    # Unset all defaults
    db.query(AIModelConfig).update({"is_default": False})
    
    # Set this one as default
    db_model.is_default = True
    db.commit()
    db.refresh(db_model)
    
    provider_value = db_model.provider.value if hasattr(db_model.provider, 'value') else str(db_model.provider)
    return {
        "id": db_model.id,
        "name": db_model.name,
        "provider": provider_value,
        "model_name": db_model.model_name,
        "base_url": db_model.base_url,
        "is_default": db_model.is_default,
        "is_active": db_model.is_active,
        "created_at": db_model.created_at
    }

@app.put("/api/ai-models/{model_id}/set-active", response_model=AIModelConfigResponse)
async def set_active_ai_model(model_id: int, db: Session = Depends(get_db)):
    """Set active AI model (use this to control which model is currently in use)"""
    db_model = db.query(AIModelConfig).filter(AIModelConfig.id == model_id).first()
    if not db_model:
        raise HTTPException(status_code=404, detail="AI model not found")
    
    if not db_model.is_active:
        # Unset all active models (only one should be active at a time)
        db.query(AIModelConfig).update({"is_active": False})
        
        # Set this one as active
        db_model.is_active = True
        db.commit()
        db.refresh(db_model)
    
    provider_value = db_model.provider.value if hasattr(db_model.provider, 'value') else str(db_model.provider)
    return {
        "id": db_model.id,
        "name": db_model.name,
        "provider": provider_value,
        "model_name": db_model.model_name,
        "base_url": db_model.base_url,
        "is_default": db_model.is_default,
        "is_active": db_model.is_active,
        "created_at": db_model.created_at
    }

# Stock Pool endpoints
@app.get("/api/stock-pools", response_model=List[StockPoolSchema])
async def get_stock_pools(db: Session = Depends(get_db)):
    """Get all stock pools"""
    pools = db.query(StockPool).order_by(StockPool.created_at.desc()).all()
    return pools

@app.get("/api/stock-pools/{pool_id}", response_model=StockPoolSchema)
async def get_stock_pool(pool_id: int, db: Session = Depends(get_db)):
    """Get a specific stock pool"""
    pool = db.query(StockPool).filter(StockPool.id == pool_id).first()
    if not pool:
        raise HTTPException(status_code=404, detail="Stock pool not found")
    return pool

@app.post("/api/stock-pools", response_model=StockPoolSchema, status_code=status.HTTP_201_CREATED)
async def create_stock_pool(pool: StockPoolCreate, db: Session = Depends(get_db)):
    """Create a new stock pool"""
    db_pool = StockPool(**pool.model_dump())
    db.add(db_pool)
    db.commit()
    db.refresh(db_pool)
    return db_pool

@app.put("/api/stock-pools/{pool_id}", response_model=StockPoolSchema)
async def update_stock_pool(pool_id: int, pool: StockPoolUpdate, db: Session = Depends(get_db)):
    """Update a stock pool"""
    db_pool = db.query(StockPool).filter(StockPool.id == pool_id).first()
    if not db_pool:
        raise HTTPException(status_code=404, detail="Stock pool not found")
    
    update_data = pool.model_dump(exclude_unset=True)
    for field, value in update_data.items():
        setattr(db_pool, field, value)
    
    db.commit()
    db.refresh(db_pool)
    return db_pool

@app.delete("/api/stock-pools/{pool_id}", status_code=status.HTTP_204_NO_CONTENT)
async def delete_stock_pool(pool_id: int, db: Session = Depends(get_db)):
    """Delete a stock pool"""
    db_pool = db.query(StockPool).filter(StockPool.id == pool_id).first()
    if not db_pool:
        raise HTTPException(status_code=404, detail="Stock pool not found")
    
    db.delete(db_pool)
    db.commit()
    return None

# Stock search endpoints
@app.get("/api/market/stocks/search")
async def search_stocks(
    q: str = "", 
    limit: int = 50, 
    market_type: Optional[str] = None,  # 'US', 'HK', 'CN'
    db: Session = Depends(get_db)
):
    """Search stocks (from database cache, with fallback to external API)"""
    results = []
    
    # First, try to search in database
    query = db.query(StockInfo)
    
    if q:
        # Search by symbol or name
        search_term = f"%{q.upper()}%"
        query = query.filter(
            (StockInfo.symbol.like(search_term)) |
            (StockInfo.name.like(search_term))
        )
    
    # Filter by market type if provided
    if market_type:
        query = query.filter(StockInfo.market_type == market_type.upper())
    
    stocks = query.limit(limit).all()
    results = [StockInfoSchema.from_orm(s) for s in stocks]
    
    # If not enough results in database, try external API fallback
    if len(results) < limit and q:
        try:
            from openbb_service import openbb_service
            import yfinance as yf
            
            # Try to search using yfinance (for US stocks)
            if not market_type or market_type.upper() == 'US':
                try:
                    # Try direct symbol lookup
                    ticker = yf.Ticker(q.upper())
                    info = ticker.info
                    if info and 'symbol' in info:
                        # Check if already in results
                        if not any(s.symbol == info['symbol'] for s in results):
                            stock_info = StockInfo(
                                symbol=info.get('symbol', q.upper()),
                                name=info.get('longName') or info.get('shortName'),
                                exchange=info.get('exchange'),
                                market_type='US',
                                sector=info.get('sector'),
                                industry=info.get('industry'),
                                market_cap=info.get('marketCap'),
                                pe_ratio=info.get('trailingPE')
                            )
                            # Try to save to database (ignore if duplicate)
                            try:
                                db.add(stock_info)
                                db.commit()
                                db.refresh(stock_info)
                            except Exception:
                                db.rollback()
                                # If exists, get from database
                                existing = db.query(StockInfo).filter(StockInfo.symbol == stock_info.symbol).first()
                                if existing:
                                    stock_info = existing
                            
                            results.append(StockInfoSchema.from_orm(stock_info))
                            if len(results) >= limit:
                                return results[:limit]
                except Exception as e:
                    logger.debug(f"yfinance search failed for {q}: {str(e)}")
            
            # For other markets or if yfinance fails, could add other data sources here
            # For now, return what we have from database
            
        except Exception as e:
            logger.warning(f"External stock search failed: {str(e)}")
    
    return results[:limit]

@app.get("/api/market/stocks/popular")
async def get_popular_stocks(limit: int = 50, market_type: Optional[str] = None, db: Session = Depends(get_db)):
    """Get popular stocks (sorted by market cap or trading volume)"""
    try:
        # Check if StockInfo table exists
        from sqlalchemy import inspect
        inspector = inspect(db.bind)
        if 'stock_info' not in inspector.get_table_names():
            # Table doesn't exist, return default stocks
            logger.warning("stock_info table does not exist, returning default stocks")
            common_stocks = [
                {'symbol': 'AAPL', 'name': 'Apple Inc.', 'market_type': 'US'},
                {'symbol': 'MSFT', 'name': 'Microsoft Corporation', 'market_type': 'US'},
                {'symbol': 'GOOGL', 'name': 'Alphabet Inc.', 'market_type': 'US'},
                {'symbol': 'AMZN', 'name': 'Amazon.com Inc.', 'market_type': 'US'},
                {'symbol': 'TSLA', 'name': 'Tesla Inc.', 'market_type': 'US'},
                {'symbol': 'META', 'name': 'Meta Platforms Inc.', 'market_type': 'US'},
                {'symbol': 'NVDA', 'name': 'NVIDIA Corporation', 'market_type': 'US'},
                {'symbol': 'JPM', 'name': 'JPMorgan Chase & Co.', 'market_type': 'US'},
                {'symbol': 'V', 'name': 'Visa Inc.', 'market_type': 'US'},
                {'symbol': 'JNJ', 'name': 'Johnson & Johnson', 'market_type': 'US'},
            ]
            if market_type:
                common_stocks = [s for s in common_stocks if s['market_type'] == market_type.upper()]
            return common_stocks[:limit]
        
        query = db.query(StockInfo)
        
        if market_type:
            query = query.filter(StockInfo.market_type == market_type.upper())
        
        # Order by market_cap descending (if available), or by symbol
        # Note: This is a simplified version - in production, you'd want to order by actual trading volume
        stocks = query.order_by(StockInfo.symbol.asc()).limit(limit).all()
        
        # If database has few stocks, fallback to common stocks
        if len(stocks) < limit:
            common_stocks = ['AAPL', 'MSFT', 'GOOGL', 'AMZN', 'TSLA', 'META', 'NVDA', 'JPM', 'V', 'JNJ']
            for symbol in common_stocks:
                if len(stocks) >= limit:
                    break
                existing = any(s.symbol == symbol for s in stocks)
                if not existing:
                    # Try to fetch from yfinance
                    try:
                        import yfinance as yf
                        ticker = yf.Ticker(symbol)
                        info = ticker.info
                        stock = StockInfo(
                            symbol=symbol,
                            name=info.get('longName', symbol),
                            market_type='US',
                            exchange=info.get('exchange', 'NASDAQ'),
                            market_cap=info.get('marketCap', 0)
                        )
                        db.add(stock)
                        stocks.append(stock)
                    except Exception as e:
                        logger.warning(f"Failed to fetch stock info for {symbol}: {e}")
                        pass
            
            try:
                db.commit()
            except Exception as e:
                logger.warning(f"Failed to commit stock info: {e}")
                db.rollback()
        
        # Convert to dict format for response
        from schemas import StockInfoSchema
        result = []
        for stock in stocks[:limit]:
            try:
                result.append(StockInfoSchema.from_orm(stock))
            except Exception as e:
                logger.warning(f"Failed to serialize stock {stock.symbol}: {e}")
                # Fallback to dict
                result.append({
                    'symbol': stock.symbol,
                    'name': getattr(stock, 'name', stock.symbol),
                    'market_type': getattr(stock, 'market_type', 'US'),
                    'exchange': getattr(stock, 'exchange', None),
                    'market_cap': getattr(stock, 'market_cap', 0)
                })
        
        return result
        
    except Exception as e:
        logger.error(f"Failed to get popular stocks: {str(e)}", exc_info=True)
        # Return default stocks on error
        common_stocks = [
            {'symbol': 'AAPL', 'name': 'Apple Inc.', 'market_type': 'US'},
            {'symbol': 'MSFT', 'name': 'Microsoft Corporation', 'market_type': 'US'},
            {'symbol': 'GOOGL', 'name': 'Alphabet Inc.', 'market_type': 'US'},
            {'symbol': 'AMZN', 'name': 'Amazon.com Inc.', 'market_type': 'US'},
            {'symbol': 'TSLA', 'name': 'Tesla Inc.', 'market_type': 'US'},
        ]
        if market_type:
            common_stocks = [s for s in common_stocks if s['market_type'] == market_type.upper()]
        return common_stocks[:limit]

@app.get("/api/market/stocks/{symbol}/info", response_model=StockInfoSchema)
async def get_stock_info(symbol: str, db: Session = Depends(get_db)):
    """Get stock detailed information"""
    stock = db.query(StockInfo).filter(StockInfo.symbol == symbol.upper()).first()
    if not stock:
        raise HTTPException(status_code=404, detail="Stock info not found")
    return stock

# Data sync endpoints (admin)
@app.post("/api/admin/sync-data", response_model=DataSyncResponse)
async def trigger_data_sync(request: DataSyncRequest, db: Session = Depends(get_db)):
    """Manually trigger data synchronization (admin)"""
    from services.data_service import DataService
    
    try:
        records_added = 0
        symbols_processed = 0
        
        async with DataService(db=db) as data_service:
            for symbol in request.symbols:
                try:
                    data = await data_service.get_historical_data(
                        symbol=symbol,
                        start_date=request.start_date,
                        end_date=request.end_date,
                        use_cache=False  # Force fetch from API
                    )
                    
                    if data is not None and not data.empty:
                        records_added += len(data)
                        symbols_processed += 1
                except Exception as e:
                    logger.error(f"Failed to sync {symbol}: {e}")
                    continue
        
        return DataSyncResponse(
            success=True,
            message=f"Synced {symbols_processed} symbols, added {records_added} records",
            symbols_processed=symbols_processed,
            records_added=records_added
        )
    except Exception as e:
        logger.error(f"Data sync failed: {e}")
        raise HTTPException(status_code=500, detail=f"Data sync failed: {str(e)}")

# Rate limit monitoring endpoints (admin)
@app.get("/api/admin/rate-limit-status")
async def get_rate_limit_status():
    """Get rate limit status (admin monitoring)"""
    try:
        from services.rate_limiter import rate_limiter
        status = rate_limiter.get_status()
        return status
    except Exception as e:
        logger.error(f"Failed to get rate limit status: {e}")
        raise HTTPException(status_code=500, detail=f"Failed to get rate limit status: {str(e)}")

@app.post("/api/admin/reset-daily-limit")
async def reset_daily_limit():
    """Reset daily request count (admin, emergency use only)"""
    try:
        from services.rate_limiter import rate_limiter
        rate_limiter.reset_daily_count()
        return {"message": "Daily limit reset successfully", "status": "ok"}
    except Exception as e:
        logger.error(f"Failed to reset daily limit: {e}")
        raise HTTPException(status_code=500, detail=f"Failed to reset daily limit: {str(e)}")

# Chat Strategy endpoints (策略提取与保存)
@app.post("/api/ai/conversations/{conversation_id}/extract-strategies", response_model=List[ChatStrategySchema])
async def extract_strategies(
    conversation_id: str,
    message_id: int = Query(..., description="Message ID to extract strategy from"),
    db: Session = Depends(get_db)
):
    """从指定消息中提取策略代码（自动识别）"""
    try:
        from services.strategy_extraction import auto_extract_strategies_from_message
        
        # 获取消息
        message = db.query(ConversationMessage).filter(
            ConversationMessage.id == message_id,
            ConversationMessage.conversation_id == conversation_id
        ).first()
        
        if not message:
            raise HTTPException(status_code=404, detail="Message not found")
        
        if message.role != 'assistant':
            raise HTTPException(status_code=400, detail="Can only extract strategies from assistant messages")
        
        # 使用策略提取服务自动提取
        extracted_strategies = auto_extract_strategies_from_message(
            message.content,
            message.code_snippets
        )
        
        if not extracted_strategies:
            raise HTTPException(status_code=400, detail="No strategy code found in message")
        
        # 为每个提取的策略创建ChatStrategy记录
        chat_strategies = []
        for strategy_info in extracted_strategies:
            # 检查是否已经提取过相同的策略（基于代码内容）
            existing = db.query(ChatStrategy).filter(
                ChatStrategy.conversation_id == conversation_id,
                ChatStrategy.message_id == message_id,
                ChatStrategy.logic_code == strategy_info['code']
            ).first()
            
            if existing:
                chat_strategies.append(existing)
                continue
            
            chat_strategy = ChatStrategy(
                conversation_id=conversation_id,
                message_id=message_id,
                name=strategy_info['name'],
                logic_code=strategy_info['code'],
                description=strategy_info.get('description')
            )
            db.add(chat_strategy)
            chat_strategies.append(chat_strategy)
        
        db.commit()
        
        # Refresh all strategies
        for cs in chat_strategies:
            db.refresh(cs)
        
        return chat_strategies
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Failed to extract strategy: {str(e)}", exc_info=True)
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Failed to extract strategy: {str(e)}")

@app.get("/api/ai/conversations/{conversation_id}/strategies", response_model=List[ChatStrategySchema])
async def get_chat_strategies(conversation_id: str, db: Session = Depends(get_db)):
    """获取会话中提取的所有策略"""
    try:
        strategies = db.query(ChatStrategy).filter(
            ChatStrategy.conversation_id == conversation_id
        ).order_by(ChatStrategy.created_at.desc()).all()
        return strategies
    except Exception as e:
        logger.error(f"Failed to get chat strategies: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to get chat strategies: {str(e)}")

@app.post("/api/ai/chat-strategies/{chat_strategy_id}/save", response_model=StrategySchema)
async def save_chat_strategy(
    chat_strategy_id: int,
    request: SaveStrategyRequest,
    db: Session = Depends(get_db)
):
    """保存聊天中的策略到策略池"""
    try:
        # 获取聊天策略
        chat_strategy = db.query(ChatStrategy).filter(
            ChatStrategy.id == chat_strategy_id
        ).first()
        
        if not chat_strategy:
            raise HTTPException(status_code=404, detail="Chat strategy not found")
        
        if chat_strategy.is_saved:
            raise HTTPException(status_code=400, detail="Strategy already saved")
        
        # 创建策略记录
        db_strategy = Strategy(
            name=request.name,
            logic_code=chat_strategy.logic_code,
            description=request.description or chat_strategy.description,
            target_portfolio_id=request.target_portfolio_id,
            is_active=False  # 默认不活跃，用户需要手动激活
        )
        db.add(db_strategy)
        db.commit()
        db.refresh(db_strategy)
        
        # 更新ChatStrategy记录
        chat_strategy.is_saved = True
        chat_strategy.saved_strategy_id = db_strategy.id
        db.commit()
        
        return db_strategy
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Failed to save chat strategy: {str(e)}")
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Failed to save chat strategy: {str(e)}")

@app.delete("/api/ai/chat-strategies/{chat_strategy_id}", status_code=status.HTTP_204_NO_CONTENT)
async def delete_chat_strategy(chat_strategy_id: int, db: Session = Depends(get_db)):
    """删除聊天中的策略"""
    try:
        chat_strategy = db.query(ChatStrategy).filter(
            ChatStrategy.id == chat_strategy_id
        ).first()
        
        if not chat_strategy:
            raise HTTPException(status_code=404, detail="Chat strategy not found")
        
        # 如果已保存到策略池，只删除ChatStrategy记录，保留Strategy记录
        # 用户可以从策略管理页面删除Strategy
        
        db.delete(chat_strategy)
        db.commit()
        return None
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Failed to delete chat strategy: {str(e)}")
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Failed to delete chat strategy: {str(e)}")

# Benchmark Strategies endpoint
@app.get("/api/backtest/benchmark-strategies")
async def get_benchmark_strategies():
    """Get list of available benchmark strategies for comparison"""
    try:
        from services.benchmark_strategies import list_benchmark_strategies
        strategies = list_benchmark_strategies()
        return strategies
    except Exception as e:
        logger.error(f"Failed to get benchmark strategies: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to get benchmark strategies: {str(e)}")

# Backtest Symbol List endpoints (回测标的清单)
@app.get("/api/backtest/symbol-lists", response_model=List[SymbolList])
async def get_symbol_lists(
    is_active: Optional[bool] = None,
    db: Session = Depends(get_db)
):
    """获取所有回测标的清单"""
    try:
        query = db.query(BacktestSymbolList)
        if is_active is not None:
            query = query.filter(BacktestSymbolList.is_active == is_active)
        lists = query.order_by(BacktestSymbolList.created_at.desc()).all()
        return lists
    except Exception as e:
        logger.error(f"Failed to get symbol lists: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to get symbol lists: {str(e)}")

@app.get("/api/backtest/symbol-lists/{list_id}", response_model=SymbolList)
async def get_symbol_list(list_id: int, db: Session = Depends(get_db)):
    """获取特定清单"""
    try:
        symbol_list = db.query(BacktestSymbolList).filter(
            BacktestSymbolList.id == list_id
        ).first()
        
        if not symbol_list:
            raise HTTPException(status_code=404, detail="Symbol list not found")
        
        return symbol_list
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Failed to get symbol list: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to get symbol list: {str(e)}")

@app.post("/api/backtest/symbol-lists", response_model=SymbolList, status_code=status.HTTP_201_CREATED)
async def create_symbol_list(list: SymbolListCreate, db: Session = Depends(get_db)):
    """创建回测标的清单"""
    try:
        db_list = BacktestSymbolList(**list.model_dump())
        db.add(db_list)
        db.commit()
        db.refresh(db_list)
        return db_list
    except Exception as e:
        logger.error(f"Failed to create symbol list: {str(e)}")
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Failed to create symbol list: {str(e)}")

@app.put("/api/backtest/symbol-lists/{list_id}", response_model=SymbolList)
async def update_symbol_list(
    list_id: int,
    list_update: SymbolListUpdate,
    db: Session = Depends(get_db)
):
    """更新回测标的清单"""
    try:
        db_list = db.query(BacktestSymbolList).filter(
            BacktestSymbolList.id == list_id
        ).first()
        
        if not db_list:
            raise HTTPException(status_code=404, detail="Symbol list not found")
        
        update_data = list_update.model_dump(exclude_unset=True)
        for field, value in update_data.items():
            setattr(db_list, field, value)
        
        db.commit()
        db.refresh(db_list)
        return db_list
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Failed to update symbol list: {str(e)}")
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Failed to update symbol list: {str(e)}")

@app.delete("/api/backtest/symbol-lists/{list_id}", status_code=status.HTTP_204_NO_CONTENT)
async def delete_symbol_list(list_id: int, db: Session = Depends(get_db)):
    """删除回测标的清单"""
    try:
        db_list = db.query(BacktestSymbolList).filter(
            BacktestSymbolList.id == list_id
        ).first()
        
        if not db_list:
            raise HTTPException(status_code=404, detail="Symbol list not found")
        
        db.delete(db_list)
        db.commit()
        return None
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Failed to delete symbol list: {str(e)}")
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Failed to delete symbol list: {str(e)}")

# Data Source Config endpoints
@app.get("/api/data-sources", response_model=List[DataSourceConfigResponse])
async def get_data_sources(db: Session = Depends(get_db)):
    """Get all data source configurations"""
    try:
        sources = db.query(DataSourceConfig).order_by(DataSourceConfig.priority.desc(), DataSourceConfig.name).all()
        result = []
        for source in sources:
            # Don't expose API key in response
            result.append(DataSourceConfigResponse(
                id=source.id,
                name=source.name,
                source_type=source.source_type,
                provider=source.provider,
                api_key=None,  # Never expose API key
                base_url=source.base_url,
                is_active=source.is_active,
                is_default=source.is_default,
                priority=source.priority,
                supports_markets=source.supports_markets,
                rate_limit=source.rate_limit,
                created_at=source.created_at,
                updated_at=source.updated_at
            ))
        return result
    except Exception as e:
        logger.error(f"Failed to get data sources: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to get data sources: {str(e)}")

@app.post("/api/data-sources", response_model=DataSourceConfigResponse, status_code=status.HTTP_201_CREATED)
async def create_data_source(source: DataSourceConfigCreate, db: Session = Depends(get_db)):
    """Create a new data source configuration"""
    try:
        from ai_service_factory import encrypt_api_key
        
        # Encrypt API key if provided
        encrypted_api_key = None
        if source.api_key:
            encrypted_api_key = encrypt_api_key(source.api_key)
        
        # If this is set as default, unset others
        if source.is_default:
            db.query(DataSourceConfig).update({"is_default": False})
        
        db_source = DataSourceConfig(
            name=source.name,
            source_type=source.source_type,
            provider=source.provider,
            api_key=encrypted_api_key,
            base_url=source.base_url,
            is_active=source.is_active,
            is_default=source.is_default,
            priority=source.priority,
            supports_markets=source.supports_markets,
            rate_limit=source.rate_limit
        )
        db.add(db_source)
        db.commit()
        db.refresh(db_source)
        
        return DataSourceConfigResponse(
            id=db_source.id,
            name=db_source.name,
            source_type=db_source.source_type,
            provider=db_source.provider,
            api_key=None,
            base_url=db_source.base_url,
            is_active=db_source.is_active,
            is_default=db_source.is_default,
            priority=db_source.priority,
            supports_markets=db_source.supports_markets,
            rate_limit=db_source.rate_limit,
            created_at=db_source.created_at,
            updated_at=db_source.updated_at
        )
    except Exception as e:
        logger.error(f"Failed to create data source: {str(e)}")
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Failed to create data source: {str(e)}")

@app.put("/api/data-sources/{source_id}", response_model=DataSourceConfigResponse)
async def update_data_source(source_id: int, source: DataSourceConfigUpdate, db: Session = Depends(get_db)):
    """Update data source configuration"""
    try:
        from ai_service_factory import encrypt_api_key
        
        db_source = db.query(DataSourceConfig).filter(DataSourceConfig.id == source_id).first()
        if not db_source:
            raise HTTPException(status_code=404, detail="Data source not found")
        
        update_data = source.model_dump(exclude_unset=True)
        
        # Encrypt API key if provided (and not empty)
        if "api_key" in update_data:
            if update_data["api_key"] and update_data["api_key"].strip():
                update_data["api_key"] = encrypt_api_key(update_data["api_key"].strip())
            else:
                # Empty API key, keep existing
                del update_data["api_key"]
        
        # If setting as default, unset others
        if update_data.get("is_default"):
            db.query(DataSourceConfig).filter(DataSourceConfig.id != source_id).update({"is_default": False})
        
        for field, value in update_data.items():
            setattr(db_source, field, value)
        
        db.commit()
        db.refresh(db_source)
        
        return DataSourceConfigResponse(
            id=db_source.id,
            name=db_source.name,
            source_type=db_source.source_type,
            provider=db_source.provider,
            api_key=None,
            base_url=db_source.base_url,
            is_active=db_source.is_active,
            is_default=db_source.is_default,
            priority=db_source.priority,
            supports_markets=db_source.supports_markets,
            rate_limit=db_source.rate_limit,
            created_at=db_source.created_at,
            updated_at=db_source.updated_at
        )
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Failed to update data source: {str(e)}")
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Failed to update data source: {str(e)}")

@app.delete("/api/data-sources/{source_id}", status_code=status.HTTP_204_NO_CONTENT)
async def delete_data_source(source_id: int, db: Session = Depends(get_db)):
    """Delete data source configuration"""
    try:
        db_source = db.query(DataSourceConfig).filter(DataSourceConfig.id == source_id).first()
        if not db_source:
            raise HTTPException(status_code=404, detail="Data source not found")
        
        db.delete(db_source)
        db.commit()
        return None
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Failed to delete data source: {str(e)}")
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Failed to delete data source: {str(e)}")

@app.get("/api/data-sources/status")
async def get_data_sources_status(db: Session = Depends(get_db)):
    """Get status of all active data sources (which one is currently being used)"""
    try:
        from models import DataSourceConfig
        from datetime import datetime, timedelta
        
        # Get all active data sources
        active_sources = db.query(DataSourceConfig).filter(
            DataSourceConfig.is_active == True
        ).order_by(DataSourceConfig.priority.desc(), DataSourceConfig.is_default.desc()).all()
        
        # Test each source with a quick fetch
        test_symbol = "AAPL"
        test_end_date = datetime.now().strftime('%Y-%m-%d')
        test_start_date = (datetime.now() - timedelta(days=7)).strftime('%Y-%m-%d')
        
        status_list = []
        working_source_id = None
        
        for source in active_sources:
            try:
                from services.data_service import DataService
                async with DataService(db=db) as data_service:
                    data_service.test_source_id = source.id
                    data = await data_service.get_historical_data(
                        test_symbol,
                        test_start_date,
                        test_end_date,
                        use_cache=False
                    )
                    
                    is_working = data is not None and not data.empty
                    if is_working and working_source_id is None:
                        working_source_id = source.id
                    
                    status_list.append({
                        "source_id": source.id,
                        "name": source.name,
                        "provider": source.provider,
                        "is_working": is_working,
                        "priority": source.priority,
                        "is_default": source.is_default,
                        "data_points": len(data) if is_working else 0,
                        "error": None
                    })
            except Exception as e:
                error_msg = str(e)
                logger.warning(f"Failed to test source {source.name}: {error_msg}")
                status_list.append({
                    "source_id": source.id,
                    "name": source.name,
                    "provider": source.provider,
                    "is_working": False,
                    "priority": source.priority,
                    "is_default": source.is_default,
                    "data_points": 0,
                    "error": error_msg
                })
        
        return {
            "sources": status_list,
            "working_source_id": working_source_id,
            "message": f"找到 {len([s for s in status_list if s['is_working']])} 个可用的数据源（共 {len(status_list)} 个激活的数据源）"
        }
    except Exception as e:
        logger.error(f"Failed to get data sources status: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Failed to get data sources status: {str(e)}")

@app.post("/api/data-sources/{source_id}/test")
async def test_data_source_connection(source_id: int, db: Session = Depends(get_db)):
    """Test connection to a data source"""
    try:
        from datetime import datetime, timedelta
        db_source = db.query(DataSourceConfig).filter(DataSourceConfig.id == source_id).first()
        if not db_source:
            raise HTTPException(status_code=404, detail="Data source not found")
        
        # Test by fetching a well-known symbol (e.g., AAPL)
        test_symbol = "AAPL"
        test_end_date = datetime.now().strftime('%Y-%m-%d')
        test_start_date = (datetime.now() - timedelta(days=30)).strftime('%Y-%m-%d')
        
        try:
            # Use the configured data source to fetch test data
            from services.data_service import DataService
            async with DataService(db=db) as data_service:
                # Temporarily set the source_id for testing
                data_service.test_source_id = source_id
                data = await data_service.get_historical_data(
                    test_symbol,
                    test_start_date,
                    test_end_date,
                    use_cache=False  # Don't use cache for testing
                )
                
                if data is not None and not data.empty:
                    return {
                        "success": True,
                        "message": f"连接成功。为 {test_symbol} 获取了 {len(data)} 个数据点。",
                        "data_points": len(data),
                        "symbol": test_symbol,
                        "date_range": f"{test_start_date} to {test_end_date}"
                    }
                else:
                    return {
                        "success": False,
                        "message": f"连接成功但未返回 {test_symbol} 的数据。",
                        "symbol": test_symbol
                    }
        except Exception as e:
            logger.error(f"Data source test failed: {str(e)}")
            return {
                "success": False,
                "message": f"连接测试失败: {str(e)}",
                "error": str(e)
            }
            
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Failed to test data source: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to test data source: {str(e)}")

# Backtest Record endpoints
@app.get("/api/backtest/records", response_model=List[BacktestRecord])
async def get_backtest_records(
    strategy_id: Optional[int] = None,
    limit: int = Query(50, le=200),
    offset: int = Query(0, ge=0),
    db: Session = Depends(get_db)
):
    """获取回测记录列表"""
    try:
        query = db.query(BacktestRecord)
        
        if strategy_id:
            query = query.filter(BacktestRecord.strategy_id == strategy_id)
        
        records = query.order_by(BacktestRecord.created_at.desc()).offset(offset).limit(limit).all()
        
        # Convert to dict and sanitize for JSON serialization
        from utils.json_serializer import sanitize_for_json
        from schemas import BacktestRecord as BacktestRecordSchema
        
        result = []
        for record in records:
            try:
                # Convert SQLAlchemy model to Pydantic schema
                record_dict = {
                    'id': record.id,
                    'name': record.name,
                    'strategy_id': record.strategy_id,
                    'strategy_name': record.strategy_name,
                    'start_date': record.start_date.isoformat() if record.start_date else None,
                    'end_date': record.end_date.isoformat() if record.end_date else None,
                    'initial_cash': record.initial_cash,
                    'symbols': record.symbols if isinstance(record.symbols, list) else [],
                    'sharpe_ratio': record.sharpe_ratio,
                    'sortino_ratio': record.sortino_ratio,
                    'annualized_return': record.annualized_return,
                    'max_drawdown': record.max_drawdown,
                    'win_rate': record.win_rate,
                    'total_trades': record.total_trades,
                    'total_return': record.total_return,
                    'full_result': sanitize_for_json(record.full_result) if record.full_result else None,
                    'compare_with_indices': getattr(record, 'compare_with_indices', False),
                    'compare_items': getattr(record, 'compare_items', None),
                    'created_at': record.created_at.isoformat() if record.created_at else None,
                    'updated_at': record.updated_at.isoformat() if record.updated_at else None,
                }
                # Validate with Pydantic schema
                result.append(BacktestRecordSchema(**record_dict))
            except Exception as e:
                logger.error(f"Failed to serialize backtest record {record.id}: {str(e)}")
                # Skip invalid records but continue processing others
                continue
        
        return result
    except Exception as e:
        logger.error(f"Failed to get backtest records: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Failed to get backtest records: {str(e)}")

@app.get("/api/backtest/records/{record_id}", response_model=BacktestRecord)
async def get_backtest_record(record_id: int, db: Session = Depends(get_db)):
    """获取单个回测记录"""
    record = db.query(BacktestRecord).filter(BacktestRecord.id == record_id).first()
    if not record:
        raise HTTPException(status_code=404, detail="Backtest record not found")
    return record

@app.put("/api/backtest/records/{record_id}", response_model=BacktestRecord)
async def update_backtest_record(
    record_id: int,
    update: BacktestRecordUpdate,
    db: Session = Depends(get_db)
):
    """更新回测记录（主要是更新名称）"""
    record = db.query(BacktestRecord).filter(BacktestRecord.id == record_id).first()
    if not record:
        raise HTTPException(status_code=404, detail="Backtest record not found")
    
    if update.name is not None:
        record.name = update.name
    
    db.commit()
    db.refresh(record)
    return record

@app.delete("/api/backtest/records/{record_id}", status_code=status.HTTP_204_NO_CONTENT)
async def delete_backtest_record(record_id: int, db: Session = Depends(get_db)):
    """删除回测记录"""
    record = db.query(BacktestRecord).filter(BacktestRecord.id == record_id).first()
    if not record:
        raise HTTPException(status_code=404, detail="Backtest record not found")
    
    db.delete(record)
    db.commit()
    return None

@app.get("/api/backtest/records/{record_id}/export/csv")
async def export_backtest_record_csv(record_id: int, db: Session = Depends(get_db)):
    """导出回测记录为CSV格式"""
    try:
        record = db.query(BacktestRecord).filter(BacktestRecord.id == record_id).first()
        if not record:
            raise HTTPException(status_code=404, detail="Backtest record not found")
        
        import io
        import csv
        
        # 创建CSV内容
        output = io.StringIO()
        writer = csv.writer(output)
        
        # 写入基本信息
        writer.writerow(['回测记录导出'])
        writer.writerow([])
        writer.writerow(['基本信息'])
        writer.writerow(['ID', record.id])
        writer.writerow(['名称', record.name or f"回测_{record.id}"])
        writer.writerow(['策略ID', record.strategy_id])
        writer.writerow(['策略名称', record.strategy_name])
        writer.writerow(['开始日期', record.start_date])
        writer.writerow(['结束日期', record.end_date])
        writer.writerow(['初始资金', record.initial_cash])
        writer.writerow(['股票列表', ', '.join(record.symbols)])
        writer.writerow(['创建时间', record.created_at])
        writer.writerow([])
        
        # 写入指标
        writer.writerow(['回测指标'])
        writer.writerow(['夏普比率', record.sharpe_ratio or 'N/A'])
        writer.writerow(['索提诺比率', record.sortino_ratio or 'N/A'])
        writer.writerow(['年化收益率', f"{record.annualized_return:.2f}%" if record.annualized_return else 'N/A'])
        writer.writerow(['最大回撤', f"{record.max_drawdown:.2f}%" if record.max_drawdown else 'N/A'])
        writer.writerow(['胜率', f"{record.win_rate:.2f}%" if record.win_rate else 'N/A'])
        writer.writerow(['总交易次数', record.total_trades or 0])
        writer.writerow(['总收益率', f"{record.total_return:.2f}%" if record.total_return else 'N/A'])
        writer.writerow([])
        
        # 如果有完整结果，导出详细数据
        if record.full_result:
            # 导出交易记录
            if 'trades' in record.full_result and record.full_result['trades']:
                writer.writerow(['交易记录'])
                writer.writerow(['日期', '股票', '方向', '价格', '数量', '佣金', '盈亏', '盈亏%', '触发原因'])
                for trade in record.full_result['trades']:
                    writer.writerow([
                        trade.get('date', ''),
                        trade.get('symbol', ''),
                        trade.get('side', ''),
                        trade.get('price', 0),
                        trade.get('quantity', 0),
                        trade.get('commission', 0),
                        trade.get('pnl', ''),
                        trade.get('pnl_percent', ''),
                        trade.get('trigger_reason', '')
                    ])
                writer.writerow([])
            
            # 导出权益曲线
            if 'equity_curve' in record.full_result and record.full_result['equity_curve']:
                writer.writerow(['权益曲线'])
                writer.writerow(['日期', '权益价值'])
                for point in record.full_result['equity_curve']:
                    writer.writerow([
                        point.get('date', ''),
                        point.get('value', 0)
                    ])
                writer.writerow([])
            
            # 导出按股票统计
            if 'per_stock_performance' in record.full_result and record.full_result['per_stock_performance']:
                writer.writerow(['按股票统计'])
                writer.writerow(['股票', '总交易次数', '买入次数', '卖出次数', '买入数量', '卖出数量', 
                               '最终持仓', '买入成本', '卖出收入', '佣金', '已实现盈亏', '收益率%'])
                for stock in record.full_result['per_stock_performance']:
                    writer.writerow([
                        stock.get('symbol', ''),
                        stock.get('total_trades', 0),
                        stock.get('buy_trades_count', 0),
                        stock.get('sell_trades_count', 0),
                        stock.get('total_quantity_bought', 0),
                        stock.get('total_quantity_sold', 0),
                        stock.get('final_position', 0),
                        stock.get('total_buy_cost', 0),
                        stock.get('total_sell_revenue', 0),
                        stock.get('total_commission', 0),
                        stock.get('realized_pnl', 0),
                        f"{stock.get('return_percent', 0):.2f}%" if stock.get('return_percent') else '0%'
                    ])
        
        output.seek(0)
        filename = f"backtest_{record_id}_{record.start_date}_{record.end_date}.csv"
        
        return StreamingResponse(
            iter([output.getvalue()]),
            media_type="text/csv; charset=utf-8-sig",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Failed to export CSV: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to export CSV: {str(e)}")

@app.get("/api/backtest/records/{record_id}/export/excel")
async def export_backtest_record_excel(record_id: int, db: Session = Depends(get_db)):
    """导出回测记录为Excel格式"""
    try:
        # 检查是否安装了openpyxl
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.utils import get_column_letter
        except ImportError:
            raise HTTPException(
                status_code=500, 
                detail="Excel export requires openpyxl. Install with: pip install openpyxl"
            )
        
        record = db.query(BacktestRecord).filter(BacktestRecord.id == record_id).first()
        if not record:
            raise HTTPException(status_code=404, detail="Backtest record not found")
        
        import io
        
        # 创建Excel工作簿
        wb = openpyxl.Workbook()
        
        # 基本信息工作表
        ws_info = wb.active
        ws_info.title = "基本信息"
        
        # 标题样式
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        # 写入基本信息
        ws_info.append(['回测记录导出'])
        ws_info.merge_cells('A1:B1')
        ws_info['A1'].font = Font(bold=True, size=14)
        
        row = 3
        info_data = [
            ['ID', record.id],
            ['名称', record.name or f"回测_{record.id}"],
            ['策略ID', record.strategy_id],
            ['策略名称', record.strategy_name],
            ['开始日期', str(record.start_date)],
            ['结束日期', str(record.end_date)],
            ['初始资金', record.initial_cash],
            ['股票列表', ', '.join(record.symbols)],
            ['创建时间', str(record.created_at)]
        ]
        for key, value in info_data:
            ws_info.cell(row=row, column=1, value=key).font = Font(bold=True)
            ws_info.cell(row=row, column=2, value=value)
            row += 1
        
        row += 1
        # 写入指标
        ws_info.append([])
        row += 1
        ws_info.append(['回测指标'])
        ws_info.merge_cells(f'A{row}:B{row}')
        ws_info[f'A{row}'].fill = header_fill
        ws_info[f'A{row}'].font = header_font
        row += 1
        
        metrics_data = [
            ['夏普比率', record.sharpe_ratio],
            ['索提诺比率', record.sortino_ratio],
            ['年化收益率', f"{record.annualized_return:.2f}%" if record.annualized_return else 'N/A'],
            ['最大回撤', f"{record.max_drawdown:.2f}%" if record.max_drawdown else 'N/A'],
            ['胜率', f"{record.win_rate:.2f}%" if record.win_rate else 'N/A'],
            ['总交易次数', record.total_trades or 0],
            ['总收益率', f"{record.total_return:.2f}%" if record.total_return else 'N/A']
        ]
        for key, value in metrics_data:
            ws_info.cell(row=row, column=1, value=key).font = Font(bold=True)
            ws_info.cell(row=row, column=2, value=value)
            row += 1
        
        # 自动调整列宽
        ws_info.column_dimensions['A'].width = 15
        ws_info.column_dimensions['B'].width = 30
        
        # 如果有完整结果，创建详细工作表
        if record.full_result:
            # 交易记录工作表
            if 'trades' in record.full_result and record.full_result['trades']:
                ws_trades = wb.create_sheet("交易记录")
                headers = ['日期', '股票', '方向', '价格', '数量', '佣金', '盈亏', '盈亏%', '触发原因']
                ws_trades.append(headers)
                
                # 设置标题样式
                for cell in ws_trades[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center')
                
                for trade in record.full_result['trades']:
                    ws_trades.append([
                        trade.get('date', ''),
                        trade.get('symbol', ''),
                        trade.get('side', ''),
                        trade.get('price', 0),
                        trade.get('quantity', 0),
                        trade.get('commission', 0),
                        trade.get('pnl', ''),
                        trade.get('pnl_percent', ''),
                        trade.get('trigger_reason', '')
                    ])
                
                # 自动调整列宽
                for column in ws_trades.columns:
                    max_length = 0
                    column_letter = get_column_letter(column[0].column)
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    ws_trades.column_dimensions[column_letter].width = min(max_length + 2, 50)
            
            # 权益曲线工作表
            if 'equity_curve' in record.full_result and record.full_result['equity_curve']:
                ws_equity = wb.create_sheet("权益曲线")
                ws_equity.append(['日期', '权益价值'])
                for cell in ws_equity[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center')
                
                for point in record.full_result['equity_curve']:
                    ws_equity.append([
                        point.get('date', ''),
                        point.get('value', 0)
                    ])
            
            # 按股票统计工作表
            if 'per_stock_performance' in record.full_result and record.full_result['per_stock_performance']:
                ws_stocks = wb.create_sheet("按股票统计")
                headers = ['股票', '总交易次数', '买入次数', '卖出次数', '买入数量', '卖出数量', 
                          '最终持仓', '买入成本', '卖出收入', '佣金', '已实现盈亏', '收益率%']
                ws_stocks.append(headers)
                
                for cell in ws_stocks[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center')
                
                for stock in record.full_result['per_stock_performance']:
                    ws_stocks.append([
                        stock.get('symbol', ''),
                        stock.get('total_trades', 0),
                        stock.get('buy_trades_count', 0),
                        stock.get('sell_trades_count', 0),
                        stock.get('total_quantity_bought', 0),
                        stock.get('total_quantity_sold', 0),
                        stock.get('final_position', 0),
                        stock.get('total_buy_cost', 0),
                        stock.get('total_sell_revenue', 0),
                        stock.get('total_commission', 0),
                        stock.get('realized_pnl', 0),
                        f"{stock.get('return_percent', 0):.2f}%" if stock.get('return_percent') else '0%'
                    ])
                
                # 自动调整列宽
                for column in ws_stocks.columns:
                    max_length = 0
                    column_letter = get_column_letter(column[0].column)
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    ws_stocks.column_dimensions[column_letter].width = min(max_length + 2, 20)
        
        # 保存到内存
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f"backtest_{record_id}_{record.start_date}_{record.end_date}.xlsx"
        
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Failed to export Excel: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to export Excel: {str(e)}")

@app.get("/api/data-sources/available")
async def get_available_data_sources():
    """Get list of available data sources (for reference)"""
    return {
        "sources": [
            {
                "name": "OpenBB Terminal",
                "provider": "openbb",
                "source_type": "free",
                "description": "Free open-source financial data platform",
                "supports_markets": ["US", "HK", "CN"],
                "rate_limit": 300,
                "requires_api_key": False
            },
            {
                "name": "Yahoo Finance",
                "provider": "yfinance",
                "source_type": "free",
                "description": "Free market data via yfinance library",
                "supports_markets": ["US", "HK"],
                "rate_limit": 2000,
                "requires_api_key": False
            },
            {
                "name": "Alpha Vantage",
                "provider": "alphavantage",
                "source_type": "free",
                "description": "Free tier: 5 API calls/minute, 500 calls/day",
                "supports_markets": ["US"],
                "rate_limit": 5,
                "requires_api_key": True,
                "api_key_url": "https://www.alphavantage.co/support/#api-key"
            },
            {
                "name": "Polygon.io",
                "provider": "polygon",
                "source_type": "paid",
                "description": "Professional market data API (paid plans available)",
                "supports_markets": ["US"],
                "rate_limit": 5,
                "requires_api_key": True,
                "api_key_url": "https://polygon.io/"
            },
            {
                "name": "IEX Cloud",
                "provider": "iexcloud",
                "source_type": "paid",
                "description": "Financial data API with free tier",
                "supports_markets": ["US"],
                "rate_limit": 100,
                "requires_api_key": True,
                "api_key_url": "https://iexcloud.io/"
            },
            {
                "name": "Twelve Data",
                "provider": "twelvedata",
                "source_type": "paid",
                "description": "Market data API with free tier",
                "supports_markets": ["US", "HK", "CN"],
                "rate_limit": 8,
                "requires_api_key": True,
                "api_key_url": "https://twelvedata.com/"
            },
            {
                "name": "Quandl/Nasdaq Data Link",
                "provider": "quandl",
                "source_type": "paid",
                "description": "Financial and economic data",
                "supports_markets": ["US", "HK", "CN"],
                "rate_limit": 50,
                "requires_api_key": True,
                "api_key_url": "https://data.nasdaq.com/"
            },
            {
                "name": "富途牛牛 (Futu)",
                "provider": "futu",
                "source_type": "free",
                "description": "富途牛牛OpenAPI，支持港股、美股、A股。需要安装OpenD客户端并登录富途账户。提供资金流向等高级数据。",
                "supports_markets": ["US", "HK", "CN"],
                "rate_limit": None,
                "requires_api_key": False,
                "requires_opend": True,
                "api_doc_url": "https://openapi.futunn.com/",
                "note": "需要安装OpenD客户端并登录富途账户"
            }
        ]
    }
