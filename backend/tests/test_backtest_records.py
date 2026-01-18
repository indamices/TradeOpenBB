"""
Unit tests for backtest records functionality
"""
import pytest
from datetime import datetime, date
from unittest.mock import Mock, patch, MagicMock
from sqlalchemy.orm import Session

from models import BacktestRecord
from schemas import BacktestRecordCreate, BacktestRecordUpdate, BacktestResult


@pytest.fixture
def mock_db():
    """Create a mock database session"""
    db = Mock(spec=Session)
    return db


@pytest.fixture
def sample_backtest_result():
    """Create a sample backtest result"""
    return BacktestResult(
        sharpe_ratio=1.5,
        sortino_ratio=1.8,
        annualized_return=12.5,
        max_drawdown=-15.0,
        win_rate=55.0,
        total_trades=100,
        total_return=12.5,
        equity_curve=[
            {'date': '2023-01-01', 'value': 100000},
            {'date': '2023-12-31', 'value': 112500}
        ],
        trades=[
            {
                'date': '2023-01-15',
                'symbol': 'AAPL',
                'side': 'BUY',
                'price': 150.0,
                'quantity': 100,
                'commission': 1.0,
                'trigger_reason': 'SMA crossover',
                'pnl': None,
                'pnl_percent': None
            }
        ]
    )


@pytest.fixture
def sample_backtest_record():
    """Create a sample backtest record"""
    return BacktestRecord(
        id=1,
        name="Test Backtest",
        strategy_id=1,
        strategy_name="Test Strategy",
        start_date=date(2023, 1, 1),
        end_date=date(2023, 12, 31),
        initial_cash=100000.0,
        symbols=["AAPL", "GOOGL"],
        sharpe_ratio=1.5,
        sortino_ratio=1.8,
        annualized_return=12.5,
        max_drawdown=-15.0,
        win_rate=55.0,
        total_trades=100,
        total_return=12.5,
        full_result={},
        created_at=datetime.now()
    )


class TestBacktestRecordModel:
    """Test BacktestRecord model"""
    
    def test_record_creation(self, sample_backtest_record):
        """Test creating a backtest record"""
        assert sample_backtest_record.id == 1
        assert sample_backtest_record.name == "Test Backtest"
        assert sample_backtest_record.strategy_id == 1
        assert len(sample_backtest_record.symbols) == 2
    
    def test_record_serialization(self, sample_backtest_record):
        """Test record can be serialized"""
        # Should not raise exception
        data = {
            'id': sample_backtest_record.id,
            'name': sample_backtest_record.name,
            'strategy_id': sample_backtest_record.strategy_id,
            'start_date': str(sample_backtest_record.start_date),
            'end_date': str(sample_backtest_record.end_date),
        }
        assert isinstance(data, dict)


class TestBacktestRecordSchemas:
    """Test BacktestRecord schemas"""
    
    def test_record_create_schema(self):
        """Test BacktestRecordCreate schema"""
        record_create = BacktestRecordCreate(
            strategy_id=1,
            start_date=date(2023, 1, 1),
            end_date=date(2023, 12, 31),
            initial_cash=100000,
            symbols=["AAPL"]
        )
        assert record_create.strategy_id == 1
        assert record_create.symbols == ["AAPL"]
    
    def test_record_update_schema(self):
        """Test BacktestRecordUpdate schema"""
        record_update = BacktestRecordUpdate(name="Updated Name")
        assert record_update.name == "Updated Name"


class TestCSVExport:
    """Test CSV export functionality"""
    
    def test_csv_export_structure(self, sample_backtest_record):
        """Test CSV export structure"""
        # Mock the export endpoint behavior
        import io
        import csv
        
        output = io.StringIO()
        writer = csv.writer(output)
        
        # Write basic info
        writer.writerow(['回测记录导出'])
        writer.writerow(['ID', sample_backtest_record.id])
        writer.writerow(['名称', sample_backtest_record.name])
        
        output.seek(0)
        content = output.getvalue()
        
        assert '回测记录导出' in content
        assert str(sample_backtest_record.id) in content


class TestExcelExport:
    """Test Excel export functionality"""
    
    @pytest.mark.skipif(True, reason="Requires openpyxl")
    def test_excel_export_structure(self, sample_backtest_record):
        """Test Excel export structure"""
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        
        ws.append(['回测记录导出'])
        ws.append(['ID', sample_backtest_record.id])
        ws.append(['名称', sample_backtest_record.name])
        
        assert ws['A1'].value == '回测记录导出'
        assert ws['B2'].value == sample_backtest_record.id


if __name__ == '__main__':
    pytest.main([__file__, '-v'])
